"""
Продвинутые фичи для бота (опционально)

Раскомментируй и добавь в main.py те функции, которые тебе нужны
"""

import asyncio
from datetime import datetime, timedelta
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes


# ============= КАТЕГОРИИ/ТЕГИ =============

async def add_category_to_word(update, context, word_id: int, category: str):
    """
    Добавить категорию к слову
    
    Использование: /category_add [word_id] [категория]
    """
    # Нужно добавить поле category в БД
    # db.add_category(word_id, category)
    pass


async def show_words_by_category(update, context, user_id: int, category: str):
    """Показать все слова определенной категории"""
    # words = db.get_words_by_category(user_id, category)
    pass


# ============= НАПОМИНАНИЯ =============

async def setup_daily_reminder(context: ContextTypes.DEFAULT_TYPE):
    """
    Ежедневное напоминание для повторения слов
    
    Добавь в main.py:
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
    
    scheduler = AsyncIOScheduler()
    scheduler.add_job(send_daily_reminder, 'cron', hour=9, minute=0)
    scheduler.start()
    """
    pass


async def send_daily_reminder(context: ContextTypes.DEFAULT_TYPE):
    """Отправить напоминание всем пользователям"""
    # Получить всех активных пользователей
    # Для каждого отправить случайное слово
    pass


# ============= ЭКСПОРТ/ИМПОРТ =============

async def export_to_pdf(update, context, user_id: int):
    """
    Экспорт словаря в PDF
    
    Установка: pip install fpdf2
    """
    from fpdf import FPDF
    
    db = Database()
    words = db.get_user_words(user_id)
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Arial', 'B', 16)
    pdf.cell(0, 10, 'Мой словарь', 0, 1, 'C')
    
    pdf.set_font('Arial', '', 12)
    for word_data in words:
        pdf.cell(0, 10, f"{word_data['word']}", 0, 1)
        pdf.multi_cell(0, 10, word_data['definition'])
        pdf.ln(5)
    
    filename = f"vocabulary_{user_id}.pdf"
    pdf.output(filename)
    
    # Отправить файл пользователю
    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=open(filename, 'rb'),
        filename='my_vocabulary.pdf'
    )


async def export_to_excel(update, context, user_id: int):
    """
    Экспорт в Excel
    
    Установка: pip install openpyxl
    """
    from openpyxl import Workbook
    
    db = Database()
    words = db.get_user_words(user_id)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Словарь"
    
    # Заголовки
    ws.append(['Слово', 'Определение', 'Дата добавления'])
    
    # Данные
    for word_data in words:
        ws.append([
            word_data['word'],
            word_data['definition'],
            word_data['created_at']
        ])
    
    filename = f"vocabulary_{user_id}.xlsx"
    wb.save(filename)
    
    await context.bot.send_document(
        chat_id=update.effective_chat.id,
        document=open(filename, 'rb'),
        filename='my_vocabulary.xlsx'
    )


# ============= КВИЗ-РЕЖИМ =============

class QuizManager:
    """Менеджер для квиз-игры"""
    
    def __init__(self, db):
        self.db = db
        self.active_quizzes = {}
    
    async def start_quiz(self, update, context, user_id: int):
        """Начать квиз"""
        words = self.db.get_user_words(user_id)
        
        if len(words) < 4:
            await update.message.reply_text(
                "Недостаточно слов для квиза. Добавь хотя бы 4 слова!"
            )
            return
        
        # Создаем квиз
        quiz_data = self._generate_quiz_question(words)
        self.active_quizzes[user_id] = {
            'current_question': 0,
            'score': 0,
            'total_questions': 5,
            'question_data': quiz_data
        }
        
        await self._send_question(update, context, user_id)
    
    def _generate_quiz_question(self, words):
        """Генерирует вопрос для квиза"""
        import random
        word_data = random.choice(words)
        
        # Варианты ответов
        other_words = [w for w in words if w['word'] != word_data['word']]
        wrong_answers = random.sample(other_words, min(3, len(other_words)))
        
        options = [word_data['word']] + [w['word'] for w in wrong_answers]
        random.shuffle(options)
        
        return {
            'definition': word_data['definition'],
            'correct_answer': word_data['word'],
            'options': options
        }
    
    async def _send_question(self, update, context, user_id: int):
        """Отправить вопрос"""
        quiz = self.active_quizzes[user_id]
        question = quiz['question_data']
        
        keyboard = []
        for option in question['options']:
            keyboard.append([
                InlineKeyboardButton(
                    option, 
                    callback_data=f"quiz_answer_{option}"
                )
            ])
        
        reply_markup = InlineKeyboardMarkup(keyboard)
        
        text = f"""
🎯 Вопрос {quiz['current_question'] + 1}/{quiz['total_questions']}

Какое слово подходит к этому определению?

"{question['definition']}"
"""
        
        await update.message.reply_text(text, reply_markup=reply_markup)


# ============= ГОЛОСОВОЙ ВВОД =============

async def handle_voice(update, context):
    """
    Обработка голосовых сообщений
    
    Установка: pip install SpeechRecognition pydub
    """
    # Скачать голосовое сообщение
    voice_file = await update.message.voice.get_file()
    await voice_file.download_to_drive('voice.ogg')
    
    # Конвертировать и распознать
    # ... (нужна дополнительная логика)
    
    await update.message.reply_text("Распознал: [слово]")


# ============= ГЕЙМИФИКАЦИЯ =============

class AchievementSystem:
    """Система достижений"""
    
    ACHIEVEMENTS = {
        'first_word': {
            'title': '🌱 Первый шаг',
            'description': 'Добавил первое слово',
            'condition': lambda stats: stats['total_words'] >= 1
        },
        'word_collector': {
            'title': '📚 Коллекционер',
            'description': 'Собрал 50 слов',
            'condition': lambda stats: stats['total_words'] >= 50
        },
        'word_master': {
            'title': '🎓 Мастер слов',
            'description': 'Собрал 100 слов',
            'condition': lambda stats: stats['total_words'] >= 100
        },
        'weekly_streak': {
            'title': '🔥 Неделя силы',
            'description': 'Добавлял слова 7 дней подряд',
            'condition': lambda stats: False  # Нужна доп. логика
        }
    }
    
    def check_achievements(self, user_id: int):
        """Проверить достижения пользователя"""
        db = Database()
        stats = db.get_user_stats(user_id)
        
        unlocked = []
        for achievement_id, achievement in self.ACHIEVEMENTS.items():
            if achievement['condition'](stats):
                unlocked.append(achievement)
        
        return unlocked


# ============= СТАТИСТИКА ПРОГРЕССА =============

def get_learning_progress(user_id: int):
    """
    Получить детальный прогресс обучения
    """
    db = Database()
    words = db.get_user_words(user_id)
    
    # Анализ по датам
    words_by_date = {}
    for word in words:
        date = word['created_at'].split()[0]  # Только дата
        words_by_date[date] = words_by_date.get(date, 0) + 1
    
    # График прогресса (можно визуализировать)
    progress = {
        'total_words': len(words),
        'words_by_date': words_by_date,
        'average_per_day': len(words) / max(len(words_by_date), 1),
        'most_productive_day': max(words_by_date.items(), key=lambda x: x[1]) if words_by_date else None
    }
    
    return progress


# ============= ИНТЕГРАЦИЯ С NOTION =============

async def sync_with_notion(user_id: int, notion_token: str, database_id: str):
    """
    Синхронизация с Notion
    
    Установка: pip install notion-client
    """
    from notion_client import Client
    
    notion = Client(auth=notion_token)
    db = Database()
    words = db.get_user_words(user_id)
    
    for word_data in words:
        # Добавить в Notion database
        notion.pages.create(
            parent={"database_id": database_id},
            properties={
                "Слово": {"title": [{"text": {"content": word_data['word']}}]},
                "Определение": {"rich_text": [{"text": {"content": word_data['definition']}}]},
            }
        )
    
    return f"Синхронизировано {len(words)} слов с Notion"


# ============= ИСПОЛЬЗОВАНИЕ =============
"""
Чтобы использовать эти фичи:

1. Раскомментируй нужные функции
2. Добавь необходимые зависимости в requirements.txt
3. Зарегистрируй обработчики в main.py:

    application.add_handler(CommandHandler("export_pdf", export_to_pdf))
    application.add_handler(CommandHandler("quiz", quiz_manager.start_quiz))
    ...

4. Обнови схему БД если нужно (добавь поля для категорий, достижений и т.д.)
"""
